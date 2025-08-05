import os
import csv
import json
import time
import smtplib
import logging
import html
import string
import random
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders
from flask import Flask, render_template, request, flash, redirect, url_for, jsonify, session
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from threading import Thread
from datetime import datetime, date, timedelta
from dotenv import load_dotenv
from openpyxl import load_workbook
from contextlib import contextmanager
from functools import wraps
import psycopg
from psycopg.rows import dict_row
from psycopg_pool import ConnectionPool
import urllib.parse

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-change-this-in-production')

# Configuration
UPLOAD_FOLDER = 'uploads'
ATTACHMENTS_FOLDER = 'attachments'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
ALLOWED_ATTACHMENT_EXTENSIONS = {'pdf', 'doc', 'docx', 'txt', 'jpg', 'jpeg', 'png', 'gif'}
MAX_CONTENT_LENGTH = 50 * 1024 * 1024  # 50MB max file size
LOGS_FOLDER = 'logs'
EMAIL_LIMIT_PER_ACCOUNT = 15  # Limit per account

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['ATTACHMENTS_FOLDER'] = ATTACHMENTS_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# Create folders if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(ATTACHMENTS_FOLDER, exist_ok=True)
os.makedirs(LOGS_FOLDER, exist_ok=True)

# Set up logging
log_level = logging.DEBUG if app.debug else logging.INFO
logging.basicConfig(
    level=log_level,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(LOGS_FOLDER, 'email_logs.log')),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)

# PostgreSQL Configuration
def get_database_url():
    """Get database URL from environment variables"""
    database_url = os.getenv('DATABASE_URL')
    if database_url:
        # Parse the URL to handle SSL requirements for production
        if database_url.startswith('postgres://'):
            database_url = database_url.replace('postgres://', 'postgresql://', 1)
        return database_url
    else:
        # Local development configuration
        return f"postgresql://{os.getenv('DB_USER', 'postgres')}:{os.getenv('DB_PASSWORD', 'password')}@{os.getenv('DB_HOST', 'localhost')}:{os.getenv('DB_PORT', '5432')}/{os.getenv('DB_NAME', 'email_system')}"

# Create connection pool
try:
    connection_pool = ConnectionPool(
        get_database_url(),
        min_size=1,
        max_size=20
    )
    if connection_pool:
        logger.info("PostgreSQL connection pool created successfully")
except Exception as e:
    logger.error(f"Error creating connection pool: {e}")
    connection_pool = None

# Database context manager
@contextmanager
def get_db_connection():
    """Get database connection from pool"""
    if not connection_pool:
        raise Exception("Database connection pool not available")
    
    with connection_pool.connection() as conn:
        yield conn

# Authentication decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Initialize database
def init_database():
    """Initialize PostgreSQL database with required tables"""
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                
                # Create users table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS users (
                        id SERIAL PRIMARY KEY,
                        email VARCHAR(255) UNIQUE NOT NULL,
                        password_hash VARCHAR(255) NOT NULL,
                        full_name VARCHAR(255) NOT NULL,
                        is_active BOOLEAN DEFAULT TRUE,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        last_login TIMESTAMP
                    )
                ''')
                
                # Create email accounts table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS email_accounts (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                        email VARCHAR(255) NOT NULL,
                        password TEXT NOT NULL,
                        is_active BOOLEAN DEFAULT TRUE,
                        sent_count INTEGER DEFAULT 0,
                        last_reset DATE DEFAULT CURRENT_DATE,
                        default_cc TEXT DEFAULT '',
                        default_bcc TEXT DEFAULT '',
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE(user_id, email)
                    )
                ''')
                
                # Create email logs table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS email_logs (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                        log_filename VARCHAR(255) NOT NULL,
                        sender_email VARCHAR(255),
                        sender_mode VARCHAR(50),
                        total_emails INTEGER,
                        sent_count INTEGER,
                        failed_count INTEGER,
                        duration_seconds REAL,
                        subject TEXT,
                        cc_emails TEXT,
                        bcc_emails TEXT,
                        attachment_name VARCHAR(255),
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        log_data JSONB
                    )
                ''')
                
                # Create email status table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS email_status (
                        id SERIAL PRIMARY KEY,
                        log_id INTEGER REFERENCES email_logs(id) ON DELETE CASCADE,
                        recipient_email VARCHAR(255),
                        sender_email VARCHAR(255),
                        status VARCHAR(50),
                        error_message TEXT,
                        sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                ''')
                
                # Create password reset OTP table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS password_reset_otp (
                        id SERIAL PRIMARY KEY,
                        email VARCHAR(255) NOT NULL,
                        otp VARCHAR(10) NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        expires_at TIMESTAMP NOT NULL,
                        used BOOLEAN DEFAULT FALSE
                    )
                ''')
                
                # Create indexes for better performance
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_email_accounts_user_id ON email_accounts(user_id)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_email_accounts_email ON email_accounts(email)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_email_logs_user_id ON email_logs(user_id)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_email_status_log_id ON email_status(log_id)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_password_reset_email ON password_reset_otp(email)')
                
                conn.commit()
                logger.info("PostgreSQL database initialized successfully")
                
    except Exception as e:
        logger.error(f"Error initializing database: {e}")
        raise e

# Initialize database on startup
try:
    init_database()
except Exception as e:
    logger.error(f"Failed to initialize database: {e}")

# Global variable to store current email sending status (per user)
email_status = {}

def get_user_email_status(user_id):
    """Get or create email status for a specific user"""
    if user_id not in email_status:
        email_status[user_id] = {
            'is_sending': False,
            'total_emails': 0,
            'sent_count': 0,
            'failed_count': 0,
            'current_email': '',
            'current_sender': '',
            'sender_rotation': {},
            'start_time': None,
            'failed_emails': [],
            'success_emails': [],
            'attachment_name': None,
            'cc_emails': [],
            'bcc_emails': []
        }
    return email_status[user_id]

# User management functions
def create_user(email, password, full_name):
    """Create a new user"""
    try:
        password_hash = generate_password_hash(password)
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('''
                    INSERT INTO users (email, password_hash, full_name)
                    VALUES (%s, %s, %s)
                    RETURNING id
                ''', (email, password_hash, full_name))
                user_id = cursor.fetchone()[0]
                conn.commit()
                logger.info(f"Created user: {email}")
                return True, user_id
    except psycopg2.IntegrityError:
        return False, "Email address already exists"
    except Exception as e:
        logger.error(f"Error creating user: {e}")
        return False, f"Error creating user: {str(e)}"

def authenticate_user(email, password):
    """Authenticate user login"""
    try:
        with get_db_connection() as conn:
            with conn.cursor(row_factory=dict_row) as cursor:
                cursor.execute('''
                    SELECT id, email, password_hash, full_name, is_active
                    FROM users 
                    WHERE email = %s AND is_active = TRUE
                ''', (email,))
                user = cursor.fetchone()
                
                if user and check_password_hash(user['password_hash'], password):
                    # Update last login
                    cursor.execute('''
                        UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE id = %s
                    ''', (user['id'],))
                    conn.commit()
                    return True, dict(user)
                return False, "Invalid email or password"
    except Exception as e:
        logger.error(f"Error authenticating user: {e}")
        return False, "Authentication error"

def get_user_by_id(user_id):
    """Get user by ID"""
    with get_db_connection() as conn:
        with conn.cursor(row_factory=dict_row) as cursor:
            cursor.execute('''
                SELECT id, email, full_name, created_at, last_login
                FROM users 
                WHERE id = %s AND is_active = TRUE
            ''', (user_id,))
            return cursor.fetchone()

# Database functions (updated with PostgreSQL syntax)
def get_email_accounts(user_id):
    """Get all email accounts for a specific user"""
    with get_db_connection() as conn:
        with conn.cursor(row_factory=dict_row) as cursor:
            cursor.execute('''
                SELECT id, email, password, is_active, sent_count, last_reset, 
                       default_cc, default_bcc, created_at 
                FROM email_accounts 
                WHERE user_id = %s AND is_active = TRUE 
                ORDER BY created_at
            ''', (user_id,))
            return cursor.fetchall()

def get_all_email_accounts(user_id):
    """Get all email accounts for a user including inactive ones"""
    with get_db_connection() as conn:
        with conn.cursor(row_factory=dict_row) as cursor:
            cursor.execute('''
                SELECT id, email, password, is_active, sent_count, last_reset, 
                       default_cc, default_bcc, created_at 
                FROM email_accounts 
                WHERE user_id = %s
                ORDER BY created_at
            ''', (user_id,))
            return cursor.fetchall()

def add_email_account(user_id, email, password, default_cc='', default_bcc=''):
    """Add new email account to database for a specific user"""
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('''
                    INSERT INTO email_accounts (user_id, email, password, default_cc, default_bcc) 
                    VALUES (%s, %s, %s, %s, %s)
                ''', (user_id, email, password, default_cc, default_bcc))
                conn.commit()
                logger.info(f"Added email account: {email} for user {user_id}")
                return True, "Email account added successfully"
    except psycopg2.IntegrityError:
        return False, "Email account already exists for this user"
    except Exception as e:
        logger.error(f"Error adding email account: {e}")
        return False, f"Error adding email account: {str(e)}"

def update_email_account(user_id, account_id, email, password, is_active, default_cc='', default_bcc=''):
    """Update email account for a specific user"""
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('''
                    UPDATE email_accounts 
                    SET email = %s, password = %s, is_active = %s, default_cc = %s, default_bcc = %s, 
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s AND user_id = %s
                ''', (email, password, is_active, default_cc, default_bcc, account_id, user_id))
                
                if cursor.rowcount > 0:
                    conn.commit()
                    logger.info(f"Updated email account ID: {account_id} for user {user_id}")
                    return True, "Email account updated successfully"
                else:
                    return False, "Email account not found or access denied"
    except psycopg2.IntegrityError:
        return False, "Email address already exists"
    except Exception as e:
        logger.error(f"Error updating email account: {e}")
        return False, f"Error updating email account: {str(e)}"

def delete_email_account(user_id, account_id):
    """Delete email account for a specific user"""
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('DELETE FROM email_accounts WHERE id = %s AND user_id = %s', (account_id, user_id))
                
                if cursor.rowcount > 0:
                    conn.commit()
                    logger.info(f"Deleted email account ID: {account_id} for user {user_id}")
                    return True, "Email account deleted successfully"
                else:
                    return False, "Email account not found or access denied"
    except Exception as e:
        logger.error(f"Error deleting email account: {e}")
        return False, f"Error deleting email account: {str(e)}"

def get_account_default_cc_bcc(user_id, sender_email):
    """Get default CC and BCC for a specific account and user"""
    with get_db_connection() as conn:
        with conn.cursor(row_factory=dict_row) as cursor:
            cursor.execute('''
                SELECT default_cc, default_bcc 
                FROM email_accounts 
                WHERE user_id = %s AND email = %s AND is_active = TRUE
            ''', (user_id, sender_email))
            result = cursor.fetchone()
            if result:
                return result['default_cc'] or '', result['default_bcc'] or ''
            return '', ''

def reset_daily_counts(user_id):
    """Reset email counts daily for a specific user"""
    today = date.today()
    
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute('''
                UPDATE email_accounts 
                SET sent_count = 0, last_reset = %s 
                WHERE user_id = %s AND last_reset < %s
            ''', (today, user_id, today))
            
            if cursor.rowcount > 0:
                conn.commit()
                logger.info(f"Reset email count for {cursor.rowcount} accounts for user {user_id}")

def get_available_sender(user_id):
    """Get next available sender account for a specific user"""
    reset_daily_counts(user_id)
    
    with get_db_connection() as conn:
        with conn.cursor(row_factory=dict_row) as cursor:
            cursor.execute('''
                SELECT email, password, default_cc, default_bcc FROM email_accounts 
                WHERE user_id = %s AND is_active = TRUE AND sent_count < %s 
                ORDER BY sent_count ASC, created_at ASC 
                LIMIT 1
            ''', (user_id, EMAIL_LIMIT_PER_ACCOUNT))
            
            result = cursor.fetchone()
            if result:
                return result['email'], result['password'], result['default_cc'], result['default_bcc']
            
            # If all accounts have reached the limit, return the first active one
            cursor.execute('''
                SELECT email, password, default_cc, default_bcc FROM email_accounts 
                WHERE user_id = %s AND is_active = TRUE 
                ORDER BY created_at ASC 
                LIMIT 1
            ''', (user_id,))
            
            result = cursor.fetchone()
            if result:
                logger.warning(f"All accounts have reached daily limit for user {user_id}. Using {result['email']}")
                return result['email'], result['password'], result['default_cc'], result['default_bcc']
    
    return None, None, None, None

def get_account_stats(user_id):
    """Get statistics for all email accounts for a specific user"""
    reset_daily_counts(user_id)
    
    with get_db_connection() as conn:
        with conn.cursor(row_factory=dict_row) as cursor:
            cursor.execute('''
                SELECT id, email, sent_count, is_active, default_cc, default_bcc, created_at
                FROM email_accounts 
                WHERE user_id = %s
                ORDER BY created_at
            ''', (user_id,))
            
            accounts = cursor.fetchall()
            stats = []
            
            for account in accounts:
                remaining = EMAIL_LIMIT_PER_ACCOUNT - account['sent_count'] if account['is_active'] else 0
                stats.append({
                    'id': account['id'],
                    'email': account['email'],
                    'sent_today': account['sent_count'],
                    'remaining': remaining,
                    'limit': EMAIL_LIMIT_PER_ACCOUNT,
                    'percentage_used': (account['sent_count'] / EMAIL_LIMIT_PER_ACCOUNT) * 100,
                    'is_active': account['is_active'],
                    'default_cc': account['default_cc'] or '',
                    'default_bcc': account['default_bcc'] or '',
                    'created_at': account['created_at']
                })
            
            return stats

def save_email_log(user_id, log_data, log_filename):
    """Save email log to database for a specific user"""
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                
                # Insert main log entry
                cursor.execute('''
                    INSERT INTO email_logs 
                    (user_id, log_filename, sender_email, sender_mode, total_emails, sent_count, 
                     failed_count, duration_seconds, subject, cc_emails, bcc_emails, 
                     attachment_name, log_data)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                ''', (
                    user_id,
                    log_filename,
                    log_data.get('sender_email'),
                    log_data.get('sender_mode', 'auto'),
                    log_data.get('total_emails', 0),
                    log_data.get('sent_count', 0),
                    log_data.get('failed_count', 0),
                    log_data.get('duration_seconds', 0),
                    log_data.get('subject', ''),
                    ', '.join(log_data.get('cc_emails', [])),
                    ', '.join(log_data.get('bcc_emails', [])),
                    log_data.get('attachment_name', ''),
                    json.dumps(log_data)
                ))
                
                log_id = cursor.fetchone()[0]
                
                # Insert individual email statuses
                for email in log_data.get('success_emails', []):
                    cursor.execute('''
                        INSERT INTO email_status (log_id, recipient_email, sender_email, status)
                        VALUES (%s, %s, %s, %s)
                    ''', (log_id, email, log_data.get('sender_email'), 'success'))
                
                for failed_entry in log_data.get('failed_emails', []):
                    if ':' in failed_entry:
                        email_addr, error = failed_entry.split(':', 1)
                        cursor.execute('''
                            INSERT INTO email_status (log_id, recipient_email, sender_email, status, error_message)
                            VALUES (%s, %s, %s, %s, %s)
                        ''', (log_id, email_addr.strip(), log_data.get('sender_email'), 'failed', error.strip()))
                
                conn.commit()
                logger.info(f"Saved email log to database: {log_filename} for user {user_id}")
                
    except Exception as e:
        logger.error(f"Error saving email log to database: {e}")

class SimpleDataFrame:
    """Simple DataFrame-like class to replace pandas"""
    def __init__(self, data, columns):
        self.data = data
        self.columns = columns
    
    def head(self, n=5):
        return SimpleDataFrame(self.data[:n], self.columns)
    
    def to_dict(self, orient='records'):
        return self.data
    
    def iterrows(self):
        for i, row in enumerate(self.data):
            yield i, row
    
    def __len__(self):
        return len(self.data)
    
    def __getitem__(self, column):
        return [row.get(column, '') for row in self.data]

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_attachment_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_ATTACHMENT_EXTENSIONS

def read_csv_file(filepath):
    """Read CSV file"""
    try:
        data = []
        with open(filepath, 'r', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            columns = list(reader.fieldnames)
            for row in reader:
                data.append(dict(row))
        return SimpleDataFrame(data, columns)
    except Exception as e:
        logger.error(f"Error reading CSV file: {e}")
        return None

def read_excel_file(filepath):
    """Read Excel file using openpyxl"""
    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active
        
        # Get column names from first row
        columns = [str(cell.value) for cell in sheet[1] if cell.value is not None]
        
        # Get data from remaining rows
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(columns):
                    row_dict[columns[i]] = str(value) if value is not None else ""
            # Only add row if it has some data
            if any(value.strip() for value in row_dict.values() if value):
                data.append(row_dict)
        
        return SimpleDataFrame(data, columns)
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        return None

def read_file(filepath):
    """Read CSV or Excel file and return SimpleDataFrame"""
    try:
        if filepath.endswith('.csv'):
            return read_csv_file(filepath)
        elif filepath.endswith(('.xlsx', '.xls')):
            return read_excel_file(filepath)
        else:
            return None
    except Exception as e:
        logger.error(f"Error reading file: {e}")
        return None

def format_email_content(content):
    """Convert plain text to HTML while preserving formatting exactly as typed"""
    # Escape HTML characters first
    content = html.escape(content)
    
    # Normalize line endings
    content = content.replace('\r\n', '\n').replace('\r', '\n')
    
    # Convert single line breaks to <br> and double line breaks to <br><br>
    # First, replace double line breaks with a unique placeholder
    content = content.replace('\n\n', '|||DOUBLE_BREAK|||')
    
    # Then replace single line breaks with <br>
    content = content.replace('\n', '<br>')
    
    # Finally, replace the placeholder with double <br>
    content = content.replace('|||DOUBLE_BREAK|||', '<br><br>')
    
    # Handle spacing
    content = content.replace('  ', '&nbsp;&nbsp;')
    content = content.replace('\t', '&nbsp;&nbsp;&nbsp;&nbsp;')
    
    # Wrap in a div with proper styling
    html_content = f'''
    <div style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
                font-size: 14px; 
                line-height: 1.2; 
                color: #333333;
                margin: 0;
                padding: 0;">
        {content}
    </div>
    '''
    
    return html_content

def parse_email_list(email_string):
    """Parse comma-separated email addresses"""
    if not email_string:
        return []
    
    emails = [email.strip() for email in email_string.split(',') if email.strip()]
    return emails

def merge_cc_bcc_lists(form_cc, form_bcc, default_cc, default_bcc):
    """Merge form CC/BCC with default CC/BCC, removing duplicates"""
    # Parse all email lists
    form_cc_list = parse_email_list(form_cc) if form_cc else []
    form_bcc_list = parse_email_list(form_bcc) if form_bcc else []
    default_cc_list = parse_email_list(default_cc) if default_cc else []
    default_bcc_list = parse_email_list(default_bcc) if default_bcc else []
    
    # Merge and remove duplicates while preserving order
    merged_cc = list(dict.fromkeys(form_cc_list + default_cc_list))
    merged_bcc = list(dict.fromkeys(form_bcc_list + default_bcc_list))
    
    return merged_cc, merged_bcc

def send_email_smtp(smtp_server, smtp_port, sender_email, sender_password, recipient_email, subject, body, user_id, cc_emails=None, bcc_emails=None, attachment_path=None):
    """Send individual email via SMTP with CC, BCC, and attachment support"""
    try:
        msg = MIMEMultipart('alternative')
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = subject
        
        # Add CC recipients
        if cc_emails:
            msg['CC'] = ', '.join(cc_emails)
        
        # BCC is handled in sendmail, not in headers
        
        # Create plain text version (keep original formatting)
        text_part = MIMEText(body, 'plain', 'utf-8')
        
        # Create HTML version with preserved formatting
        formatted_body = format_email_content(body)
        html_part = MIMEText(formatted_body, 'html', 'utf-8')
        
        # Attach parts to message (order matters - plain text first, then HTML)
        msg.attach(text_part)
        msg.attach(html_part)
        
        # Add attachment if provided
        if attachment_path and os.path.exists(attachment_path):
            try:
                with open(attachment_path, 'rb') as f:
                    attachment_data = f.read()
                
                filename = os.path.basename(attachment_path)
                
                # Determine MIME type based on file extension
                if filename.lower().endswith('.pdf'):
                    attachment = MIMEApplication(attachment_data, _subtype='pdf')
                else:
                    attachment = MIMEBase('application', 'octet-stream')
                    attachment.set_payload(attachment_data)
                    encoders.encode_base64(attachment)
                
                attachment.add_header('Content-Disposition', f'attachment; filename={filename}')
                msg.attach(attachment)
                logger.info(f"Attached file: {filename}")
                
            except Exception as e:
                logger.warning(f"Failed to attach file {attachment_path}: {e}")
        
        # Prepare recipient list (To + CC + BCC)
        all_recipients = [recipient_email]
        if cc_emails:
            all_recipients.extend(cc_emails)
        if bcc_emails:
            all_recipients.extend(bcc_emails)
        
        # Send email
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        text = msg.as_string()
        server.sendmail(sender_email, all_recipients, text)
        server.quit()
        
        # Update the account's sent count in database
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('''
                    UPDATE email_accounts 
                    SET sent_count = sent_count + 1 
                    WHERE email = %s AND user_id = %s
                ''', (sender_email, user_id))
                conn.commit()
        
        logger.info(f"Email sent successfully to {recipient_email} from {sender_email} for user {user_id}")
        return True, "Email sent successfully"
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Failed to send email to {recipient_email} from {sender_email}: {error_msg}")
        return False, error_msg

# OTP management functions (updated for PostgreSQL)
def generate_otp():
    """Generate a 6-digit OTP"""
    return ''.join(random.choices(string.digits, k=6))

def save_otp(email, otp):
    """Save OTP to database with expiration"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Clean up old OTPs for this email
            cursor.execute('DELETE FROM password_reset_otp WHERE email = %s', (email,))
            
            # Calculate expiration time (10 minutes from now)
            expires_at = datetime.now() + timedelta(minutes=10)
            
            # Insert new OTP
            cursor.execute('''
                INSERT INTO password_reset_otp (email, otp, expires_at)
                VALUES (%s, %s, %s)
            ''', (email, otp, expires_at))
            
            conn.commit()
            logger.info(f"OTP saved for email: {email}")

def verify_otp(email, otp):
    """Verify OTP and mark as used"""
    with get_db_connection() as conn:
        with conn.cursor(row_factory=dict_row) as cursor:
            cursor.execute('''
                SELECT id, expires_at FROM password_reset_otp 
                WHERE email = %s AND otp = %s AND used = FALSE
                ORDER BY created_at DESC LIMIT 1
            ''', (email, otp))
            
            result = cursor.fetchone()
            if not result:
                return False, "Invalid OTP"
            
            # Check if OTP has expired
            if datetime.now() > result['expires_at']:
                return False, "OTP has expired"
            
            # Mark OTP as used
            cursor.execute('UPDATE password_reset_otp SET used = TRUE WHERE id = %s', (result['id'],))
            conn.commit()
            
            return True, "OTP verified successfully"

def send_otp_email(recipient_email, otp):
    """Send OTP via email"""
    try:
        otp_sender_email = os.getenv('OTP_EMAIL')
        otp_sender_password = os.getenv('OTP_EMAIL_PASSWORD')
        
        if not otp_sender_email or not otp_sender_password:
            logger.error("OTP email credentials not configured in environment variables")
            return False, "Email service not configured"
        
        subject = "Password Reset OTP - Bulk Email Sender"
        
        # Create HTML email content
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background-color: #007bff; color: white; padding: 20px; text-align: center; border-radius: 8px 8px 0 0; }}
                .content {{ background-color: #f8f9fa; padding: 30px; border-radius: 0 0 8px 8px; }}
                .otp {{ font-size: 32px; font-weight: bold; color: #007bff; text-align: center; letter-spacing: 8px; margin: 20px 0; padding: 15px; background-color: white; border: 2px dashed #007bff; border-radius: 8px; }}
                .warning {{ background-color: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; border-radius: 8px; margin: 20px 0; }}
                .footer {{ text-align: center; margin-top: 30px; color: #666; font-size: 14px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🔐 Password Reset Request</h1>
                </div>
                <div class="content">
                    <h2>Hello!</h2>
                    <p>You requested to reset your password for your Bulk Email Sender account.</p>
                    <p>Your One-Time Password (OTP) is:</p>
                    
                    <div class="otp">{otp}</div>
                    
                    <div class="warning">
                        <strong>⚠️ Important:</strong>
                        <ul>
                            <li>This OTP is valid for <strong>10 minutes only</strong></li>
                            <li>Do not share this OTP with anyone</li>
                            <li>If you didn't request this, please ignore this email</li>
                        </ul>
                    </div>
                    
                    <p>Enter this OTP on the password reset page to create your new password.</p>
                    
                    <p>Best regards,<br>
                    Bulk Email Sender Team</p>
                </div>
                <div class="footer">
                    <p>This is an automated message. Please do not reply to this email.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Create plain text version
        text_content = f"""
        Password Reset Request - Bulk Email Sender
        
        Hello!
        
        You requested to reset your password for your Bulk Email Sender account.
        
        Your One-Time Password (OTP) is: {otp}
        
        IMPORTANT:
        - This OTP is valid for 10 minutes only
        - Do not share this OTP with anyone
        - If you didn't request this, please ignore this email
        
        Enter this OTP on the password reset page to create your new password.
        
        Best regards,
        Bulk Email Sender Team
        
        This is an automated message. Please do not reply to this email.
        """
        
        # Send email using the same SMTP function
        msg = MIMEMultipart('alternative')
        msg['From'] = otp_sender_email
        msg['To'] = recipient_email
        msg['Subject'] = subject
        
        # Add both text and HTML parts
        text_part = MIMEText(text_content, 'plain', 'utf-8')
        html_part = MIMEText(html_content, 'html', 'utf-8')
        
        msg.attach(text_part)
        msg.attach(html_part)
        
        # Send email
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(otp_sender_email, otp_sender_password)
        text = msg.as_string()
        server.sendmail(otp_sender_email, [recipient_email], text)
        server.quit()
        
        logger.info(f"OTP email sent successfully to {recipient_email}")
        return True, "OTP email sent successfully"
        
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Failed to send OTP email to {recipient_email}: {error_msg}")
        return False, f"Failed to send email: {error_msg}"

def reset_user_password(email, new_password):
    """Reset user password"""
    try:
        password_hash = generate_password_hash(new_password)
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('''
                    UPDATE users SET password_hash = %s WHERE email = %s AND is_active = TRUE
                ''', (password_hash, email))
                
                if cursor.rowcount > 0:
                    conn.commit()
                    logger.info(f"Password reset successfully for user: {email}")
                    return True, "Password reset successfully"
                else:
                    return False, "User not found"
    except Exception as e:
        logger.error(f"Error resetting password: {e}")
        return False, f"Error resetting password: {str(e)}"

# Authentication Routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        
        if not email or not password:
            flash('Please enter both email and password.', 'error')
            return render_template('auth/login.html')
        
        success, result = authenticate_user(email, password)
        if success:
            session['user_id'] = result['id']
            session['email'] = result['email']
            session['full_name'] = result['full_name']
            flash(f'Welcome back, {result["full_name"]}!', 'success')
            return redirect(url_for('index'))
        else:
            flash(result, 'error')
    
    return render_template('auth/login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        full_name = request.form.get('full_name', '').strip()
        
        # Validation
        if not all([email, password, confirm_password, full_name]):
            flash('All fields are required.', 'error')
            return render_template('auth/signup.html')
        
        if password != confirm_password:
            flash('Passwords do not match.', 'error')
            return render_template('auth/signup.html')
        
        if len(password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return render_template('auth/signup.html')
        
        success, result = create_user(email, password, full_name)
        if success:
            flash('Account created successfully! Please log in.', 'success')
            return redirect(url_for('login'))
        else:
            flash(result, 'error')
    
    return render_template('auth/signup.html')

@app.route('/logout')
def logout():
    user_email = session.get('email', 'User')
    session.clear()
    flash(f'You have been logged out successfully, {user_email}!', 'info')
    return redirect(url_for('login'))

@app.route('/profile')
@login_required
def profile():
    user = get_user_by_id(session['user_id'])
    account_stats = get_account_stats(session['user_id'])
    
    # Get total logs count
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) as count FROM email_logs WHERE user_id = %s', (session['user_id'],))
        total_logs = cursor.fetchone()['count']
    
    return render_template('auth/profile.html', user=user, account_stats=account_stats, total_logs=total_logs)

# Main Routes (protected)
@app.route('/')
@login_required
def index():
    account_stats = get_account_stats(session['user_id'])
    return render_template('index.html', account_stats=account_stats)

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        flash('No file selected')
        return redirect(url_for('index'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected')
        return redirect(url_for('index'))
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Read file to get columns
        df = read_file(filepath)
        if df is not None:
            columns = list(df.columns)
            sample_data = df.head().to_dict('records')
            account_stats = get_account_stats(session['user_id'])
            
            return render_template('compose.html', 
                                 filename=filename, 
                                 columns=columns,
                                 sample_data=sample_data,
                                 account_stats=account_stats)
        else:
            flash('Failed to read the uploaded file')
            return redirect(url_for('index'))
    else:
        flash('Invalid file type. Please upload CSV or Excel files only.')
        return redirect(url_for('index'))

@app.route('/upload_attachment', methods=['POST'])
@login_required
def upload_attachment():
    """Handle attachment upload"""
    if 'attachment' not in request.files:
        return jsonify({'success': False, 'message': 'No file selected'})
    
    file = request.files['attachment']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if file and allowed_attachment_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['ATTACHMENTS_FOLDER'], filename)
        file.save(filepath)
        
        # Get file size in MB
        file_size = os.path.getsize(filepath) / (1024 * 1024)
        
        return jsonify({
            'success': True, 
            'filename': filename,
            'filepath': filepath,
            'size': f"{file_size:.2f} MB"
        })
    else:
        return jsonify({'success': False, 'message': 'Invalid file type. Allowed: PDF, DOC, DOCX, TXT, JPG, PNG, GIF'})

@app.route('/send_emails', methods=['POST'])
@login_required
def send_emails():
    user_id = session['user_id']
    filename = request.form.get('filename')
    sender_mode = request.form.get('sender_mode')
    selected_sender = request.form.get('selected_sender')
    subject = request.form.get('subject')
    template = request.form.get('template')
    email_column = request.form.get('email_column')
    delay = int(request.form.get('delay', 1))
    cc_emails = request.form.get('cc_emails', '').strip()
    bcc_emails = request.form.get('bcc_emails', '').strip()
    attachment_filename = request.form.get('attachment_filename', '').strip()
    
    # Debug logging
    logger.info(f"Form data received - filename: {filename}, sender_mode: {sender_mode}, selected_sender: '{selected_sender}', subject: {subject}, template length: {len(template) if template else 0}, email_column: {email_column}, cc: {cc_emails}, bcc: {bcc_emails}, attachment: {attachment_filename}")
    
    # Validate all required fields
    if not all([filename, sender_mode, subject, template, email_column]):
        flash('All fields are required.')
        logger.error("Missing required fields")
        return redirect(url_for('index'))
    
    # Clean up selected_sender - convert empty string to None
    if selected_sender == '' or selected_sender is None:
        selected_sender = None
    
    # Validate sender selection for manual mode
    if sender_mode == 'manual':
        if not selected_sender:
            flash('Please select an email account for manual mode.')
            logger.error("Manual mode selected but no sender specified")
            return redirect(url_for('index'))
        
        # Validate that the selected sender exists in database
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT id FROM email_accounts WHERE user_id = %s AND email = %s AND is_active = 1', (user_id, selected_sender))
            if not cursor.fetchone():
                flash(f'Selected email account "{selected_sender}" is not configured or inactive.')
                logger.error(f"Selected sender {selected_sender} not found in database")
                return redirect(url_for('index'))
    
    # Check if we have any email accounts
    accounts = get_email_accounts(user_id)
    if not accounts:
        flash('No email accounts configured. Please add email accounts first.')
        logger.error("No email accounts configured")
        return redirect(url_for('manage_accounts'))
    
    # Check if selected sender has remaining quota
    if sender_mode == 'manual' and selected_sender:
        reset_daily_counts(user_id)
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT sent_count FROM email_accounts WHERE user_id = %s AND email = %s', (user_id, selected_sender))
            account = cursor.fetchone()
            if account and account['sent_count'] >= EMAIL_LIMIT_PER_ACCOUNT:
                flash(f'Selected account {selected_sender} has reached its daily limit of {EMAIL_LIMIT_PER_ACCOUNT} emails.')
                logger.warning(f"Account {selected_sender} has reached daily limit")
                return redirect(url_for('index'))
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    # Validate that the file exists
    if not os.path.exists(filepath):
        flash('Uploaded file not found. Please upload the file again.')
        logger.error(f"File not found: {filepath}")
        return redirect(url_for('index'))
    
    # Handle attachment
    attachment_path = None
    if attachment_filename:
        attachment_path = os.path.join(app.config['ATTACHMENTS_FOLDER'], attachment_filename)
        if not os.path.exists(attachment_path):
            flash('Attachment file not found. Please upload the attachment again.')
            logger.error(f"Attachment not found: {attachment_path}")
            return redirect(url_for('index'))
    
    # Start sending emails in background thread
    def send_emails_task():
        try:
            if sender_mode == 'auto':
                logger.info("Starting auto-rotate email sending")
                success, result = send_bulk_emails(
                    user_id, filepath, subject, template, email_column, delay, 
                    cc_emails, bcc_emails, attachment_path
                )
            else:  # manual mode
                logger.info(f"Starting manual email sending with sender: {selected_sender}")
                success, result = send_bulk_emails_single_sender(
                    user_id, filepath, selected_sender, subject, template, email_column, delay,
                    cc_emails, bcc_emails, attachment_path
                )
            
            if success:
                flash(f"Bulk email sending completed! {result['success_count']} sent, {result['failed_count']} failed in {result['duration']}")
                if result.get('sender_rotation'):
                    rotation_info = ", ".join([f"{email}: {count}" for email, count in result['sender_rotation'].items()])
                    flash(f"Sender distribution: {rotation_info}")
                if result['failed_emails']:
                    flash(f"Failed emails: {', '.join(result['failed_emails'][:5])}")
                logger.info(f"Bulk email sending completed successfully: {result}")
            else:
                flash(f"Error: {result}")
                logger.error(f"Bulk email sending failed: {result}")
        except Exception as e:
            error_msg = f"Error in email sending task: {str(e)}"
            flash(error_msg)
            logger.error(error_msg, exc_info=True)
    
    thread = Thread(target=send_emails_task)
    thread.daemon = True
    thread.start()
    
    flash('Email sending started in background. Check the status page for real-time updates.')
    logger.info("Email sending thread started successfully")
    return redirect(url_for('status'))

@app.route('/manage_accounts')
@login_required
def manage_accounts():
    """Display email account management page"""
    account_stats = get_account_stats(session['user_id'])
    return render_template('manage_accounts.html', account_stats=account_stats)

@app.route('/add_account', methods=['POST'])
@login_required
def add_account():
    """Add new email account"""
    user_id = session['user_id']
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()
    default_cc = request.form.get('default_cc', '').strip()
    default_bcc = request.form.get('default_bcc', '').strip()
    
    if not email or not password:
        flash('Email and password are required.')
        return redirect(url_for('manage_accounts'))
    
    success, message = add_email_account(user_id, email, password, default_cc, default_bcc)
    flash(message)
    return redirect(url_for('manage_accounts'))

@app.route('/update_account/<int:account_id>', methods=['POST'])
@login_required
def update_account(account_id):
    """Update email account"""
    user_id = session['user_id']
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()
    is_active = request.form.get('is_active') == 'on'
    default_cc = request.form.get('default_cc', '').strip()
    default_bcc = request.form.get('default_bcc', '').strip()
    
    if not email or not password:
        flash('Email and password are required.')
        return redirect(url_for('manage_accounts'))
    
    success, message = update_email_account(user_id, account_id, email, password, is_active, default_cc, default_bcc)
    flash(message)
    return redirect(url_for('manage_accounts'))

@app.route('/delete_account/<int:account_id>', methods=['POST'])
@login_required
def delete_account(account_id):
    """Delete email account"""
    user_id = session['user_id']
    success, message = delete_email_account(user_id, account_id)
    flash(message)
    return redirect(url_for('manage_accounts'))

@app.route('/status')
@login_required
def status():
    """Display real-time email sending status"""
    user_id = session['user_id']
    user_status = get_user_email_status(user_id)
    account_stats = get_account_stats(user_id)
    return render_template('status.html', status=user_status, account_stats=account_stats)

@app.route('/api/status')
@login_required
def api_status():
    """API endpoint for real-time status updates"""
    user_id = session['user_id']
    user_status = get_user_email_status(user_id)
    status_data = user_status.copy()
    status_data['account_stats'] = get_account_stats(user_id)
    return jsonify(status_data)

@app.route('/logs')
@login_required
def logs():
    """Display email sending logs from database for current user"""
    user_id = session['user_id']
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT el.*, 
                   COUNT(es.id) as total_recipients,
                   SUM(CASE WHEN es.status = 'success' THEN 1 ELSE 0 END) as successful_sends,
                   SUM(CASE WHEN es.status = 'failed' THEN 1 ELSE 0 END) as failed_sends
            FROM email_logs el
            LEFT JOIN email_status es ON el.id = es.log_id
            WHERE el.user_id = %s
            GROUP BY el.id
            ORDER BY el.created_at DESC
            LIMIT 50
        ''', (user_id,))
        
        logs = []
        for row in cursor.fetchall():
            log_data = {
                'id': row['id'],
                'log_filename': row['log_filename'],
                'sender_email': row['sender_email'],
                'sender_mode': row['sender_mode'],
                'total_emails': row['total_emails'],
                'sent_count': row['sent_count'],
                'failed_count': row['failed_count'],
                'duration_seconds': row['duration_seconds'],
                'subject': row['subject'],
                'cc_emails': row['cc_emails'],
                'bcc_emails': row['bcc_emails'],
                'attachment_name': row['attachment_name'],
                'created_at': row['created_at'],
                'total_recipients': row['total_recipients'] or 0,
                'successful_sends': row['successful_sends'] or 0,
                'failed_sends': row['failed_sends'] or 0
            }
            logs.append(log_data)
    
    return render_template('logs.html', logs=logs)

@app.route('/log_details/<int:log_id>')
@login_required
def log_details(log_id):
    """Display detailed log information for current user"""
    user_id = session['user_id']
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Get log info (ensure it belongs to current user)
        cursor.execute('SELECT * FROM email_logs WHERE id = %s AND user_id = %s', (log_id, user_id))
        log = cursor.fetchone()
        
        if not log:
            flash('Log not found.')
            return redirect(url_for('logs'))
        
        # Get email statuses
        cursor.execute('''
            SELECT recipient_email, sender_email, status, error_message, sent_at
            FROM email_status 
            WHERE log_id = %s
            ORDER BY sent_at
        ''', (log_id,))
        
        email_statuses = cursor.fetchall()
    
    return render_template('log_details.html', log=log, email_statuses=email_statuses)

# Health check endpoint for Render
@app.route('/health')
def health_check():
    """Health check endpoint for deployment platforms"""
    return jsonify({'status': 'healthy', 'timestamp': datetime.now().isoformat()})

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        
        if not email:
            flash('Please enter your email address.', 'error')
            return render_template('auth/forgot_password.html')
        
        # Check if user exists
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT email FROM users WHERE email = %s AND is_active = 1', (email,))
            user = cursor.fetchone()
            
            if not user:
                flash('If this email exists in our system, you will receive an OTP shortly.', 'info')
                return render_template('auth/forgot_password.html')
        
        # Generate and send OTP
        otp = generate_otp()
        save_otp(email, otp)
        
        success, message = send_otp_email(email, otp)
        if success:
            flash('OTP sent to your email address. Check your inbox.', 'success')
            return redirect(url_for('verify_otp_page', email=email))
        else:
            flash('Failed to send OTP. Please try again later.', 'error')
            logger.error(f"Failed to send OTP to {email}: {message}")
    
    return render_template('auth/forgot_password.html')

@app.route('/verify_otp')
def verify_otp_page():
    email = request.args.get('email')
    if not email:
        flash('Invalid request. Please start the password reset process again.', 'error')
        return redirect(url_for('forgot_password'))
    
    return render_template('auth/verify_otp.html', email=email)

@app.route('/verify_otp', methods=['POST'])
def verify_otp_submit():
    email = request.form.get('email', '').strip()
    otp = request.form.get('otp', '').strip()
    
    if not email or not otp:
        flash('Please enter the OTP.', 'error')
        return render_template('auth/verify_otp.html', email=email)
    
    success, message = verify_otp(email, otp)
    if success:
        # Store email in session for password reset
        session['reset_email'] = email
        session['otp_verified'] = True
        flash('OTP verified successfully. Please set your new password.', 'success')
        return redirect(url_for('reset_password'))
    else:
        flash(message, 'error')
        return render_template('auth/verify_otp.html', email=email)

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    # Check if OTP was verified
    if not session.get('otp_verified') or not session.get('reset_email'):
        flash('Please complete the OTP verification first.', 'error')
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        if not password or not confirm_password:
            flash('Please enter both password fields.', 'error')
            return render_template('auth/reset_password.html')
        
        if password != confirm_password:
            flash('Passwords do not match.', 'error')
            return render_template('auth/reset_password.html')
        
        if len(password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return render_template('auth/reset_password.html')
        
        email = session.get('reset_email')
        success, message = reset_user_password(email, password)
        
        if success:
            # Clear session data
            session.pop('reset_email', None)
            session.pop('otp_verified', None)
            flash('Password reset successfully! Please log in with your new password.', 'success')
            return redirect(url_for('login'))
        else:
            flash(message, 'error')
    
    return render_template('auth/reset_password.html')